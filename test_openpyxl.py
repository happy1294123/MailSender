import openpyxl
workbook = openpyxl.load_workbook('./users.xlsx')
worksheet = workbook.active

dictionary = {}
for row in range(1, worksheet.max_row + 1):
  key = worksheet.cell(row, 1).value
  value = worksheet.cell(row, 2).value
  dictionary[key] = value

print(dictionary)