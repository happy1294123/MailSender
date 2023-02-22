""" 查找excel相關資料 """
from openpyxl import load_workbook
import config

class ExcelFinder:
    def __init__(self):
        wb = load_workbook(config.account_list_path)
        self.account_list = wb.active
        wb = load_workbook(config.email_distribution_path)
        self.email_distribution = wb.active

    def find_account_by_fund_id(self, fund_id):
        fund_id_col = '';
        for cell in self.account_list[1]:
            if cell.value == 'age': # 改為fund_id的header
                fund_id_col = cell.column - 1
            elif cell.value == 'Account Type':
                account_type_col = cell.column - 1
            elif cell.value == 'Managed By Asset Class':
                mbac_col = cell.column - 1
            elif cell.value == 'Alpha Profolio Manager':
                apm_col = cell.column - 1

        for row in self.account_list.iter_rows(values_only=True):
            if row[fund_id_col] == fund_id:
                return {
                    'account_type': row[account_type_col],
                    'mbac': row[mbac_col],
                    'apm': row[apm_col]
                }

    def find_email_by_account(self, account):
        # 找到address ＆ note
        pass



    def find(self):
        pass



ef = ExcelFinder()
account = ef.find_account_by_fund_id(12)
print(account)


